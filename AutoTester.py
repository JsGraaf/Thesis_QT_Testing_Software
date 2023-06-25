### A program to automatically generate verilog code for the Quicklogic Thing ###
### Plus, flash it and run power tests when connected to a uC with an INA219  ###
### and the companion controller                                              ###

import argparse
import subprocess
import os
import serial
import openpyxl
import datetime
import time
import threading

VERILOG_GENERATION_FILES_PATH = r"Verilog_Generators/"
QORC_SDK_PATH = r"/home/joris/Uni/Scriptie/qorc_sdk/"
QT_APP_DIR = r"QT_App"
QT_APP_PATH = r"QT_App/fpga/rtl/"
FPGA_FILE_PATH = os.path.join(QT_APP_PATH, "MODULE_top.v")
OUTPUT_DIR = r"./Results/"

HEADERS = ["Timing(ms)", "Voltage(V)", "Amperage(mA)", "Wattage(mW)", "LDO2 Voltage(V)"]

LDO_2_TRIP_LEVEL = 0.95
QT_RESET_TIME = 6

ERROR_IN_GENERATOR_EXECUTION = 1
ERROR_IN_EXECUTING_MAKE = 2
ERROR_IN_SERIAL_CONNECTION = 3
ERROR_IN_RESETTING_BOARD = 4
ERROR_IN_FLASHING_BOARD = 5
ERROR_IN_POWER_TEST = 6

threadReturn = 0
firstRun = True
filename = ''

def parseArguments():
   argParser = argparse.ArgumentParser()
   argParser.add_argument("-q", "--QORC_PORT", required=True,
                          help="Specify the serial port to which the QT+ is connected in programming mode")
   argParser.add_argument("-p", "--COM_PORT", required=True,
                          help="Specify the serial port to which the controller uC with INA219 is connected")
   argParser.add_argument("-g", "--generator", required=True,
                          help="Specify the filename of the verilog generator, located in Verilog_Generators/")
   argParser.add_argument("-c", "--circuitCount", default=200,
                          help="Specify the amount of circuits to generate")
   argParser.add_argument("-s", "--circuitStart", default=0,
                          help="The starting amount of circuits to generate")
   argParser.add_argument("-i", "--increment", default=10,
                          help="Amount to increment the circuits by")
   argParser.add_argument("-l", "--testLength", default=50000,
                          help="Specify the length of the power tests")
   argParser.add_argument("-d", "--delay", default=500,
                          help="Specify delay between measurements")
   return argParser.parse_args()

def printError(errno: int):
   if (errno == ERROR_IN_GENERATOR_EXECUTION):
      print("Error in generator execution!")
   elif (errno == ERROR_IN_EXECUTING_MAKE):
      print("Error in executing make!")
   elif (errno == ERROR_IN_SERIAL_CONNECTION):
      print("Error in serial connection!")
   elif (errno == ERROR_IN_RESETTING_BOARD):
      print("Error in Resetting board!")
   elif (errno == ERROR_IN_FLASHING_BOARD):
      print("Error in flashing board!")
   elif (errno == ERROR_IN_POWER_TEST):
      print("Error in power test!")
   global threadReturn
   threadReturn = errno
   exit(errno)

# Runs the specified generator, located in VERILOG_GENERATION_FILES_PATH
def runGenerator(filePath: os.path, circuitCount: int):
   # Prepend './'
   filePath = os.path.join("./", filePath)
   # Run the bash generator script
   val = subprocess.check_call("{} {} {}".format(filePath, circuitCount, 
                               FPGA_FILE_PATH), shell=True, 
                               stdout=subprocess.DEVNULL)
   
   if (val == 0): return True
   else: return False

# Resets the board by pulling down RESET 
# And then by pulling down USR_BTN (GPIO6)
def enableProgrammingMode(ser: serial.Serial):
   # Test connection
   if (not checkSerConnection(ser)): return False
   ser.flush()
   # Reset board using uC connected to RST pin and GPIO6 (USR_BTN)
   # Pull down USR_BTN (GPIO6)
   ser.write(b'r')
   read = ser.read()
   if (read.decode('UTF-8') != 'c'): 
      print("QT+ reset command not received!")
      return False
   print("Waiting for QT+ to finish entering programming mode...")
   read = ser.read()
   if (read.decode('UTF-8') != 'd'):
      print("Invalid code after QT+ reset: {}".format(read))
      return False
   return True

# Compiles the QT_APP using make -C
def compileProgram():
   val = subprocess.check_call("make -C {}".format(QT_APP_DIR), shell=True, 
                               stdout=subprocess.DEVNULL)
   if (val == 0): return True
   else: return False

# Test serial connection
def checkSerConnection(ser: serial.Serial):
   ser.flush()
   ser.write(b't')
   read = ser.read()
   if (read.decode('UTF-8') != 'c'): 
      print("Serial connection fail: {}".format(read))
      return False
   read = ser.read()
   if (read.decode('UTF-8') != 'c'): 
      print("Serial connection fail: {}".format(read))
      return False
   return True

# Start Serial connection
def startSerial(COM_PORT: str):
   print("Starting Serial...")
   ser = serial.Serial(COM_PORT)
   if (ser.name != COM_PORT): return False
   # Test connection
   if (not checkSerConnection(ser)): return False
   ser.flush()
   return ser

# Flasing board over USB
# COM_PORT: port which is connected to the uC for controlling the QT+
# QORC_PORT: port to which the QT+ is connected for transferring data
def flashBoard(ser):
   print("Putting QT+ into programming mode")
   if not enableProgrammingMode(ser): 
      printError(ERROR_IN_RESETTING_BOARD)
   # Wait for COM Port to show up
   time.sleep(QT_RESET_TIME)
   # Using make flash
   print("Flashing...")
   val = subprocess.check_call("make flash -C {}".format(QT_APP_DIR), 
                               shell=True)
   if (val == 0): return True
   else: return False

# Output measurements string to csv file
# Sample seperated by ';', values by ':'
def outputToCsv(sheetName: str, measurements: list, args):
   print("Creating sheet {}".format(sheetName))
   # Check if output dir exists
   if not os.path.isdir(OUTPUT_DIR):
      os.mkdir(OUTPUT_DIR)
   # Open Workbook
   global filename
   if (filename == ''):
      filename = args.generator + "_OUTPUT.xlsx"
   workbook = None
   worksheet = None
   global firstRun
   if (firstRun):   
      while (os.path.exists(os.path.join(OUTPUT_DIR, filename))):
         filename = "new_" + filename
      workbook = openpyxl.Workbook()
      worksheet = workbook.active
      worksheet.title=sheetName
      firstRun = False
   else: 
      if (os.path.exists(os.path.join(OUTPUT_DIR, filename))):
         workbook = openpyxl.load_workbook(os.path.join(OUTPUT_DIR, filename)) 
         worksheet = workbook.create_sheet(sheetName)
      else:
         print("WARNING: Had to create a new workbook after the first iteration!")
         workbook = openpyxl.Workbook()
         worksheet = workbook.active
         worksheet.title=sheetName

   # Add headers
   for i in range(1, len(HEADERS)+1):
      worksheet.cell(row=1, column=i, value=HEADERS[i-1])
   # Add date as final header
   worksheet.cell(row=1, column=i+1, 
                  value = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
   # Add data
   rowCounter = 2 # Start at one due to header row
   # Split measurments string
   measurements = measurements.split(';')
   # Remove index containing 'END' keyword
   del measurements[-1]
   # Check if there are measurement
   print(measurements)
   # Keep track of averages
   averageV = 0
   averagemA = 0
   averagemW = 0
   LDO2Tripped = False
   for row in measurements:
      sample = row.split(':')
      averageV += float(sample[0])
      averagemA += float(sample[1])
      averagemW += float(sample[2])
      if (float(sample[3]) < LDO_2_TRIP_LEVEL): LDO2Tripped = True
      # Add timing info
      worksheet.cell(row=rowCounter, column = 1, 
                     value = int(args.delay) * (rowCounter-2))
      for column in range(2, len(HEADERS)+1):
         worksheet.cell(row=rowCounter, column=column, 
                        value=float(sample[column-2]))
      rowCounter += 1
   
   # Add average values
   worksheet.cell(row=rowCounter, column=2, value=averageV/len(measurements))
   worksheet.cell(row=rowCounter, column=3, value=averagemA/len(measurements))
   worksheet.cell(row=rowCounter, column=4, value=averagemW/len(measurements))
   worksheet.cell(row=rowCounter, column=5, value=LDO2Tripped)
   worksheet.cell(row=rowCounter, column=6, value='Average')
   # Add row count beside the date
   worksheet.cell(row=1, column=i+2, value = rowCounter-2)
   workbook.save(os.path.join(OUTPUT_DIR, filename))
   return True

# Run power tests by sending serial command to uC connected to INA219 to 
# perform power tests for the desired amount of time
# COM_PORT: Port to which the uC with the INA219 is connected
def runPowerTests(ser: serial.Serial, testLength: int, betweenTime: int, 
                  outName: str, circuitCount: int):
   # Send command to uC to confirm link
   if (not checkSerConnection(ser)): return False
   # Send command to uC to start testing for power with length and sample delay
   ser.write(bytes("p {} {}".format(testLength, betweenTime), 'UTF-8'))
   read = ser.read()
   if (read.decode('UTF-8') != 'c'): 
      print("Measurements start command not received!")
      return False
   print("Starting power tests with length: {}ms, delay: {}ms".format(
                                                                  testLength, 
                                                                  betweenTime))
   # Receive all the data and chop up into individual measurements
   measurements = ser.read_until(expected=b"END").decode("UTF-8")
   print("Done! Outputting to CSV...")
   # Output to CSV file
   outputToCsv(outName + "_" + str(circuitCount), measurements, args)
   return True

def generateAndCompile(generator, i):
   print("Generating {} with {} circuits".format(generator, i))
   if not runGenerator(os.path.join(VERILOG_GENERATION_FILES_PATH, generator), 
                       i):
      printError(ERROR_IN_GENERATOR_EXECUTION)
   
   print("Compiling...")
   if not compileProgram():
      printError(ERROR_IN_EXECUTING_MAKE)

def mainLoop(args, makeThread):
   for i in range(int(args.increment) + int(args.circuitStart), int(args.circuitCount), 
                  int(args.increment)):
      print("Running {}".format(i))
      global threadReturn
      if makeThread.isAlive():
         print("Waiting for next program to compile!")
         makeThread.join()
         if (threadReturn != 0):
            printError(threadReturn)
      print("Flasing board")
      if not flashBoard(ser):
         printError(ERROR_IN_FLASHING_BOARD)
      ### Build the next program in the background ###
      makeThread = threading.Thread(target=generateAndCompile, 
                                    args=(args.generator, i + int(args.increment)), daemon=True)
      makeThread.start()
      ### Test the currently flashed program ###
      # Wait for the FPGA program to start
      time.sleep(QT_RESET_TIME)
      if not runPowerTests(ser, args.testLength, args.delay, args.generator, i):
         printError(ERROR_IN_POWER_TEST)
   # Finish final iteration
   if makeThread.isAlive():
      print("Waiting for next program to compile!")
      makeThread.join()
      if (threadReturn != 0):
            printError(threadReturn)
   if not flashBoard(ser):
      printError(ERROR_IN_FLASHING_BOARD)

   ### Test the final program ###
   # Wait for the FPGA program to start
   time.sleep(QT_RESET_TIME)
   if not runPowerTests(ser, args.testLength, args.delay, args.generator, 
                        args.circuitCount):
      printError(ERROR_IN_POWER_TEST)

if __name__ == '__main__':
   args = parseArguments()
   # Setting the QORC_PORT in the environment
   os.environ['QORC_PORT'] = args.QORC_PORT
   # Run the first compile outside of the loop
   generateAndCompile(args.generator, args.circuitStart)
   # Start serial and flash the first program
   ser = startSerial(args.COM_PORT)
   if (ser == False):
      printError(ERROR_IN_SERIAL_CONNECTION)
   if not flashBoard(ser):
      printError(ERROR_IN_FLASHING_BOARD)
   ### Build the next program in the background ###
   if (int(args.circuitStart) + int(args.increment) < int(args.circuitCount)):
      makeThread = threading.Thread(target=generateAndCompile, 
                                    args=(args.generator, int(args.circuitStart) + int(args.increment)), 
                                    daemon=True)
      makeThread.start()
   # Wait for the FPGA program to start
   time.sleep(QT_RESET_TIME)
   ### Test the currently flashed program ###
   if not runPowerTests(ser, args.testLength, args.delay, args.generator, args.circuitStart):
      printError(ERROR_IN_POWER_TEST)
   
   if (int(args.circuitStart) + int(args.increment) < int(args.circuitCount)):
      try: 
         mainLoop(args, makeThread)
      except Exception as e:
         print("EXCEPTION: {}".format(e))
         exit(-1)